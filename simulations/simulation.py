"""
Simulation Runner v2
====================
Configurable simulation: set # players, # bots, # weeks.
Outputs Excel (PnL, fees, dividends) and PNG (price charts).

All 6 fixes integrated:
- Flaw 1: Negative events + panic selling
- Flaw 2: Agent churn (rage quit + delayed replacement)
- Flaw 3: Balanced buy/sell + rebalancer agent type
- Flaw 4: Higher initial liquidity + warm-up period excluded from reports
- Flaw 5: Passive now faces real drawdown from events
- Flaw 6: Dividend attribution (base vs outperformer breakdown)
"""

import os
import sys
import random
from typing import Dict, List

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from simulations.engine import MarketEngine, generate_players
from simulations.bots import (
    AgentState, create_agents, STRATEGIES,
    handle_panic_selling, process_churn,
)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    HAS_MATPLOTLIB = True
except ImportError:
    HAS_MATPLOTLIB = False


class SimConfig:
    def __init__(
        self,
        num_players: int = 20,
        agent_config: Dict[str, int] = None,
        starting_cash: float = 1000,
        num_weeks: int = 8,
        games_per_week: int = 2,
        trading_sessions_per_game: int = 3,
        top_n_outperformers: int = 10,
        initial_pool_shares: int = 5000,    # Flaw 4: was 1000, now 5000 for deeper liquidity
        initial_price: float = 10.0,
        fee_rate: float = 0.015,
        warmup_weeks: int = 2,              # Flaw 4: exclude first N weeks from report
        churn_rate: float = 0.06,           # Flaw 2: weekly churn probability per agent
        seed: int = None,
        output_dir: str = None,
    ):
        self.num_players = num_players
        self.agent_config = agent_config or {
            "random": 8,
            "momentum": 5,
            "value": 3,
            "passive": 5,
            "rebalancer": 4,    # Flaw 3: new type
        }
        self.starting_cash = starting_cash
        self.num_weeks = num_weeks
        self.games_per_week = games_per_week
        self.trading_sessions_per_game = trading_sessions_per_game
        self.top_n_outperformers = top_n_outperformers
        self.initial_pool_shares = initial_pool_shares
        self.initial_price = initial_price
        self.fee_rate = fee_rate
        self.warmup_weeks = warmup_weeks
        self.churn_rate = churn_rate
        self.seed = seed
        self.output_dir = output_dir or os.path.join(
            os.path.dirname(os.path.abspath(__file__)), "output"
        )


def run_simulation(config: SimConfig):
    if config.seed is not None:
        random.seed(config.seed)

    os.makedirs(config.output_dir, exist_ok=True)

    print(f"{'='*60}")
    print(f"DIVIDEND FANTASY SIMULATION v2")
    print(f"{'='*60}")
    print(f"Players:          {config.num_players}")
    print(f"Agents:           {sum(config.agent_config.values())} "
          f"({', '.join(f'{v} {k}' for k, v in config.agent_config.items())})")
    print(f"Weeks:            {config.num_weeks} ({config.warmup_weeks} warm-up)")
    print(f"Games/week:       {config.games_per_week}")
    print(f"Top-N dividends:  {config.top_n_outperformers}")
    print(f"Fee rate:         {config.fee_rate*100}%")
    print(f"Starting cash:    ${config.starting_cash}")
    print(f"Pool liquidity:   {config.initial_pool_shares} shares @ ${config.initial_price}")
    print(f"Churn rate:       {config.churn_rate*100}%/week")
    print(f"Seed:             {config.seed}")
    print(f"{'='*60}\n")

    # Initialize
    players = generate_players(config.num_players)
    engine = MarketEngine(
        players, config.initial_pool_shares, config.initial_price,
        config.fee_rate, config.top_n_outperformers
    )
    agents = create_agents(config.agent_config, config.starting_cash)
    pending_agents: List[AgentState] = []  # Flaw 2: agents waiting to join

    day = 0
    weekly_summaries = []
    total_churned = 0
    total_replaced = 0

    # Record initial state
    engine.record_prices(day)
    for agent in agents:
        agent.record_pnl(day, 0, engine)

    for week in range(1, config.num_weeks + 1):
        is_warmup = week <= config.warmup_weeks
        label = " (warm-up)" if is_warmup else ""
        print(f"Week {week}/{config.num_weeks}{label}...")

        # Activate any pending agents whose join week has arrived
        for pa in pending_agents[:]:
            if pa.joined_week <= week:
                agents.append(pa)
                pending_agents.remove(pa)
                total_replaced += 1

        for game_num in range(1, config.games_per_week + 1):
            # Pre-game trading
            for session in range(config.trading_sessions_per_game):
                day += 1
                for agent in agents:
                    if not agent.active or agent.joined_week > week:
                        continue
                    strategy = STRATEGIES[agent.agent_type]
                    strategy(agent, engine, week, game_num, day)

                engine.record_prices(day)
                for agent in agents:
                    agent.record_pnl(day, week, engine)

            # Game happens — includes negative event rolls
            new_events = engine.generate_game(week)
            day += 1

            # Flaw 1: Panic selling on negative events
            if new_events:
                event_names = [f"{engine.players[e.player_id].name} ({e.event_type})"
                               for e in new_events]
                print(f"  Events: {', '.join(event_names)}")
                handle_panic_selling(agents, engine, new_events, week, game_num, day)

            # Post-game trading
            for agent in agents:
                if not agent.active or agent.joined_week > week:
                    continue
                strategy = STRATEGIES[agent.agent_type]
                strategy(agent, engine, week, game_num, day)

            engine.record_prices(day)
            for agent in agents:
                agent.record_pnl(day, week, engine)

        # End of week: churn
        new_replacements = process_churn(agents, engine, week, 0, day, config.churn_rate)
        churned_this_week = len(new_replacements)
        total_churned += churned_this_week
        pending_agents.extend(new_replacements)

        # Count trades this week
        week_trades = sum(1 for t in engine.trades if t.week == week)

        # Distribute dividends
        agent_holdings = {
            a.id: dict(a.holdings) for a in agents if a.active
        }
        div_result = engine.distribute_dividends(week, agent_holdings)

        # Credit dividends with attribution (Flaw 6)
        for agent in agents:
            if not agent.active:
                continue
            base_div = div_result["base_dividends"].get(agent.id, 0)
            outperf_div = div_result["outperformer_dividends"].get(agent.id, 0)
            total_div = base_div + outperf_div

            agent.cash += total_div
            agent.total_dividends_received += total_div
            agent.total_base_dividends += base_div
            agent.total_outperformer_dividends += outperf_div

        # Record post-dividend
        for agent in agents:
            agent.record_pnl(day, week, engine)

        summary = {
            "week": week,
            "is_warmup": is_warmup,
            "trades": week_trades,
            "fees": div_result["fees_collected"],
            "company_revenue": div_result["company_revenue"],
            "dividend_pool": div_result["dividend_pool"],
            "base_pool": div_result["base_pool"],
            "outperformer_pool": div_result["outperformer_pool"],
            "total_distributed": div_result["total_distributed"],
            "top_outperformers": div_result["top_outperformers"],
            "churned": churned_this_week,
            "active_agents": sum(1 for a in agents if a.active),
            "events": len([e for e in engine.event_log if e.week == week]),
        }
        weekly_summaries.append(summary)

        churn_str = f" | Churned: {churned_this_week}" if churned_this_week > 0 else ""
        event_str = f" | Events: {summary['events']}" if summary['events'] > 0 else ""
        print(f"  Trades: {week_trades} | Fees: ${div_result['fees_collected']:.2f} | "
              f"Divs: ${div_result['total_distributed']:.2f} | "
              f"Active: {summary['active_agents']}{churn_str}{event_str}")

    # ── Final Report ──────────────────────────────────────────────
    print(f"\n{'='*60}")
    print(f"SIMULATION COMPLETE")
    print(f"{'='*60}")
    print(f"Total trades:       {len(engine.trades)}")
    print(f"Total fees:         ${engine.total_fees:.2f}")
    print(f"Total dividends:    ${sum(d['total_distributed'] for d in weekly_summaries):.2f}")
    print(f"Negative events:    {len(engine.event_log)}")
    print(f"Agents churned:     {total_churned}")
    print(f"Agents replaced:    {total_replaced}")

    # Count winners vs losers
    active_agents = [a for a in agents if a.active]
    winners = sum(1 for a in agents if a.portfolio_value(engine) > a.initial_cash)
    losers = sum(1 for a in agents if a.portfolio_value(engine) <= a.initial_cash)
    print(f"\nWinners: {winners} | Losers: {losers} | "
          f"Win rate: {winners/(winners+losers)*100:.0f}%")

    print(f"\nAgent Results (excluding warm-up):")
    print(f"{'Agent':<25} {'Type':<12} {'PnL':>10} {'PnL%':>8} "
          f"{'Fees':>8} {'BaseDv':>8} {'OutDv':>8} {'NetDvROI':>9} {'Status':>8}")
    print("-" * 100)

    sorted_agents = sorted(agents, key=lambda a: a.portfolio_value(engine), reverse=True)
    for agent in sorted_agents:
        pv = agent.portfolio_value(engine)
        pnl = pv - agent.initial_cash
        pnl_pct = pnl / agent.initial_cash * 100 if agent.initial_cash > 0 else 0
        net_div_roi = ((agent.total_dividends_received - agent.total_fees_paid)
                       / agent.initial_cash * 100) if agent.initial_cash > 0 else 0
        status = "active" if agent.active else f"quit_w{agent.exited_week}"

        print(f"{agent.id:<25} {agent.agent_type:<12} "
              f"${pnl:>+9.2f} {pnl_pct:>+7.1f}% "
              f"${agent.total_fees_paid:>7.2f} "
              f"${agent.total_base_dividends:>7.2f} "
              f"${agent.total_outperformer_dividends:>7.2f} "
              f"{net_div_roi:>+8.1f}% "
              f"{status:>8}")

    # By type averages
    print(f"\nAverage by Type:")
    print(f"{'Type':<12} {'Avg PnL':>10} {'Avg PnL%':>9} {'Avg Fees':>10} "
          f"{'Avg BaseDv':>11} {'Avg OutDv':>10} {'NetDvROI':>9}")
    print("-" * 75)
    for atype in config.agent_config.keys():
        type_agents = [a for a in agents if a.agent_type == atype]
        if not type_agents:
            continue
        avg_pnl = sum(a.portfolio_value(engine) - a.initial_cash for a in type_agents) / len(type_agents)
        avg_pct = avg_pnl / config.starting_cash * 100
        avg_fees = sum(a.total_fees_paid for a in type_agents) / len(type_agents)
        avg_base = sum(a.total_base_dividends for a in type_agents) / len(type_agents)
        avg_outperf = sum(a.total_outperformer_dividends for a in type_agents) / len(type_agents)
        avg_net_roi = ((avg_base + avg_outperf - avg_fees) / config.starting_cash * 100)
        print(f"{atype:<12} ${avg_pnl:>+9.2f} {avg_pct:>+8.1f}% ${avg_fees:>9.2f} "
              f"${avg_base:>10.2f} ${avg_outperf:>9.2f} {avg_net_roi:>+8.1f}%")

    # Generate outputs
    if HAS_OPENPYXL:
        excel_path = _write_excel(config, engine, agents, weekly_summaries)
        print(f"\nExcel: {excel_path}")
    else:
        print("\nInstall openpyxl for Excel output: pip install openpyxl")

    if HAS_MATPLOTLIB:
        png_path = _write_charts(config, engine, agents, weekly_summaries)
        print(f"Charts: {png_path}")
    else:
        print("Install matplotlib for chart output: pip install matplotlib")

    return engine, agents, weekly_summaries


# ── Excel Output ──────────────────────────────────────────────────

def _write_excel(config, engine, agents, weekly_summaries):
    wb = openpyxl.Workbook()

    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, size=11)
    money_fmt = '#,##0.00'
    warmup_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    def style_header(ws, max_col):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    # ── Sheet 1: Daily PnL ──
    ws = wb.active
    ws.title = "Daily PnL"

    headers = ["Day", "Week"]
    for agent in agents:
        headers.extend([
            f"{agent.id} Portfolio",
            f"{agent.id} PnL",
            f"{agent.id} PnL%",
        ])
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, len(headers))

    max_days = max(len(a.pnl_history) for a in agents)
    for row_idx in range(max_days):
        row = row_idx + 2
        if row_idx < len(agents[0].pnl_history):
            snap = agents[0].pnl_history[row_idx]
            ws.cell(row=row, column=1, value=snap["day"])
            ws.cell(row=row, column=2, value=snap["week"])

        col = 3
        for agent in agents:
            if row_idx < len(agent.pnl_history):
                snap = agent.pnl_history[row_idx]
                ws.cell(row=row, column=col, value=snap["portfolio_value"]).number_format = money_fmt
                ws.cell(row=row, column=col+1, value=snap["pnl"]).number_format = money_fmt
                ws.cell(row=row, column=col+2, value=snap["pnl_pct"]).number_format = '0.00'
            col += 3

    # ── Sheet 2: Weekly Fees & Dividends ──
    ws2 = wb.create_sheet("Weekly Fees & Dividends")
    fee_headers = ["Week", "Warm-up?", "Trades", "Total Fees", "Company Revenue",
                    "Dividend Pool", "Base Pool", "Outperformer Pool",
                    "Total Distributed", "Events", "Churned", "Active Agents"]
    for i, h in enumerate(fee_headers, 1):
        ws2.cell(row=1, column=i, value=h)
    style_header(ws2, len(fee_headers))

    for row_idx, summary in enumerate(weekly_summaries, 2):
        ws2.cell(row=row_idx, column=1, value=summary["week"])
        ws2.cell(row=row_idx, column=2, value="Yes" if summary["is_warmup"] else "No")
        ws2.cell(row=row_idx, column=3, value=summary["trades"])
        ws2.cell(row=row_idx, column=4, value=round(summary["fees"], 2)).number_format = money_fmt
        ws2.cell(row=row_idx, column=5, value=round(summary["company_revenue"], 2)).number_format = money_fmt
        ws2.cell(row=row_idx, column=6, value=round(summary["dividend_pool"], 2)).number_format = money_fmt
        ws2.cell(row=row_idx, column=7, value=round(summary["base_pool"], 2)).number_format = money_fmt
        ws2.cell(row=row_idx, column=8, value=round(summary["outperformer_pool"], 2)).number_format = money_fmt
        ws2.cell(row=row_idx, column=9, value=round(summary["total_distributed"], 2)).number_format = money_fmt
        ws2.cell(row=row_idx, column=10, value=summary["events"])
        ws2.cell(row=row_idx, column=11, value=summary["churned"])
        ws2.cell(row=row_idx, column=12, value=summary["active_agents"])

        # Highlight warm-up weeks
        if summary["is_warmup"]:
            for col in range(1, len(fee_headers) + 1):
                ws2.cell(row=row_idx, column=col).fill = warmup_fill

    # ── Sheet 3: Agent Summary (Flaw 6: full dividend attribution) ──
    ws3 = wb.create_sheet("Agent Summary")
    agent_headers = ["Agent", "Type", "Status", "Initial Cash", "Final Portfolio",
                     "PnL ($)", "PnL (%)", "Total Fees Paid",
                     "Base Dividends", "Outperformer Dividends", "Total Dividends",
                     "Net Div ROI (%)", "# Holdings"]
    for i, h in enumerate(agent_headers, 1):
        ws3.cell(row=1, column=i, value=h)
    style_header(ws3, len(agent_headers))

    sorted_agents = sorted(agents, key=lambda a: a.portfolio_value(engine), reverse=True)
    for row_idx, agent in enumerate(sorted_agents, 2):
        pv = agent.portfolio_value(engine)
        pnl = pv - agent.initial_cash
        status = "Active" if agent.active else f"Quit (week {agent.exited_week})"
        net_roi = ((agent.total_dividends_received - agent.total_fees_paid)
                   / agent.initial_cash * 100) if agent.initial_cash > 0 else 0

        ws3.cell(row=row_idx, column=1, value=agent.id)
        ws3.cell(row=row_idx, column=2, value=agent.agent_type)
        ws3.cell(row=row_idx, column=3, value=status)
        ws3.cell(row=row_idx, column=4, value=agent.initial_cash).number_format = money_fmt
        ws3.cell(row=row_idx, column=5, value=round(pv, 2)).number_format = money_fmt
        ws3.cell(row=row_idx, column=6, value=round(pnl, 2)).number_format = money_fmt
        ws3.cell(row=row_idx, column=7, value=round(pnl/agent.initial_cash*100, 2) if agent.initial_cash > 0 else 0).number_format = '0.00'
        ws3.cell(row=row_idx, column=8, value=round(agent.total_fees_paid, 2)).number_format = money_fmt
        ws3.cell(row=row_idx, column=9, value=round(agent.total_base_dividends, 2)).number_format = money_fmt
        ws3.cell(row=row_idx, column=10, value=round(agent.total_outperformer_dividends, 2)).number_format = money_fmt
        ws3.cell(row=row_idx, column=11, value=round(agent.total_dividends_received, 2)).number_format = money_fmt
        ws3.cell(row=row_idx, column=12, value=round(net_roi, 2)).number_format = '0.00'
        ws3.cell(row=row_idx, column=13, value=len(agent.holdings))

    # ── Sheet 4: Price History ──
    ws4 = wb.create_sheet("Price History")
    player_ids = list(engine.price_history.keys())
    price_headers = ["Day"] + [engine.players[pid].name for pid in player_ids]
    for i, h in enumerate(price_headers, 1):
        ws4.cell(row=1, column=i, value=h)
    style_header(ws4, len(price_headers))

    if player_ids:
        max_entries = max(len(engine.price_history[pid]) for pid in player_ids)
        for row_idx in range(max_entries):
            row = row_idx + 2
            if row_idx < len(engine.price_history[player_ids[0]]):
                ws4.cell(row=row, column=1, value=engine.price_history[player_ids[0]][row_idx][0])
            for col_idx, pid in enumerate(player_ids, 2):
                if row_idx < len(engine.price_history[pid]):
                    ws4.cell(row=row, column=col_idx,
                             value=round(engine.price_history[pid][row_idx][1], 4)).number_format = '0.0000'

    # ── Sheet 5: Trade Log ──
    ws5 = wb.create_sheet("Trade Log")
    trade_headers = ["Week", "Game", "Day", "Agent", "Player", "Side",
                     "Shares", "Cost/Revenue", "Fee", "Price Before", "Price After"]
    for i, h in enumerate(trade_headers, 1):
        ws5.cell(row=1, column=i, value=h)
    style_header(ws5, len(trade_headers))

    for row_idx, trade in enumerate(engine.trades, 2):
        ws5.cell(row=row_idx, column=1, value=trade.week)
        ws5.cell(row=row_idx, column=2, value=trade.game)
        ws5.cell(row=row_idx, column=3, value=trade.day)
        ws5.cell(row=row_idx, column=4, value=trade.agent_id)
        ws5.cell(row=row_idx, column=5, value=trade.player_id)
        ws5.cell(row=row_idx, column=6, value=trade.side)
        ws5.cell(row=row_idx, column=7, value=trade.shares)
        ws5.cell(row=row_idx, column=8, value=round(trade.cost_or_revenue, 2)).number_format = money_fmt
        ws5.cell(row=row_idx, column=9, value=round(trade.fee, 2)).number_format = money_fmt
        ws5.cell(row=row_idx, column=10, value=round(trade.price_before, 4)).number_format = '0.0000'
        ws5.cell(row=row_idx, column=11, value=round(trade.price_after, 4)).number_format = '0.0000'

    # ── Sheet 6: Negative Events Log ──
    ws6 = wb.create_sheet("Events Log")
    event_headers = ["Week", "Player", "Event Type", "Severity", "Duration (weeks)"]
    for i, h in enumerate(event_headers, 1):
        ws6.cell(row=1, column=i, value=h)
    style_header(ws6, len(event_headers))

    for row_idx, event in enumerate(engine.event_log, 2):
        ws6.cell(row=row_idx, column=1, value=event.week)
        ws6.cell(row=row_idx, column=2, value=engine.players[event.player_id].name)
        ws6.cell(row=row_idx, column=3, value=event.event_type)
        ws6.cell(row=row_idx, column=4, value=event.severity)
        ws6.cell(row=row_idx, column=5, value=event.duration)

    # Auto-width key sheets
    for ws_obj in [ws2, ws3, ws6]:
        for col in range(1, ws_obj.max_column + 1):
            max_len = max(
                len(str(ws_obj.cell(row=r, column=col).value or ""))
                for r in range(1, min(ws_obj.max_row + 1, 50))
            )
            ws_obj.column_dimensions[get_column_letter(col)].width = min(max_len + 3, 25)

    path = os.path.join(config.output_dir, "simulation_results.xlsx")
    wb.save(path)
    return path


# ── Chart Output ──────────────────────────────────────────────────

def _write_charts(config, engine, agents, weekly_summaries):
    fig, axes = plt.subplots(2, 2, figsize=(16, 12))
    fig.suptitle("Dividend Fantasy Simulation v2", fontsize=16, fontweight="bold")

    type_colors = {"random": "#e74c3c", "momentum": "#3498db",
                   "value": "#2ecc71", "passive": "#9b59b6",
                   "rebalancer": "#f39c12"}

    # ── Chart 1: Player Prices (showing both ups AND downs) ──
    ax1 = axes[0, 0]
    player_ids = list(engine.price_history.keys())
    price_volatility = {}
    for pid in player_ids:
        prices = [p for _, p in engine.price_history[pid]]
        if len(prices) > 1:
            price_volatility[pid] = max(prices) - min(prices)

    top_players = sorted(price_volatility, key=price_volatility.get, reverse=True)[:8]
    for pid in top_players:
        days = [d for d, _ in engine.price_history[pid]]
        prices = [p for _, p in engine.price_history[pid]]
        ax1.plot(days, prices, label=engine.players[pid].name, linewidth=1.5)

    ax1.axhline(y=config.initial_price, color="gray", linewidth=0.8, linestyle="--", alpha=0.5)
    # Mark warm-up boundary
    if config.warmup_weeks > 0:
        warmup_days = config.warmup_weeks * config.games_per_week * (config.trading_sessions_per_game + 1)
        ax1.axvline(x=warmup_days, color="orange", linewidth=1, linestyle="--", alpha=0.7, label="Warm-up end")

    ax1.set_title("Player Token Prices (most volatile)")
    ax1.set_xlabel("Day")
    ax1.set_ylabel("Price ($)")
    ax1.legend(fontsize=7, loc="upper left")
    ax1.grid(True, alpha=0.3)

    # ── Chart 2: Agent PnL Over Time ──
    ax2 = axes[0, 1]
    for atype in config.agent_config.keys():
        type_agents = [a for a in agents if a.agent_type == atype]
        if not type_agents:
            continue
        max_len = max(len(a.pnl_history) for a in type_agents)
        avg_pnl = []
        for i in range(max_len):
            vals = [a.pnl_history[i]["pnl"] for a in type_agents if i < len(a.pnl_history)]
            avg_pnl.append(sum(vals) / len(vals) if vals else 0)
        days = list(range(max_len))
        color = type_colors.get(atype, None)
        ax2.plot(days, avg_pnl, label=f"{atype} (avg)", linewidth=2, color=color)

    ax2.axhline(y=0, color="black", linewidth=0.8, linestyle="--")
    ax2.set_title("Average PnL by Agent Type")
    ax2.set_xlabel("Day")
    ax2.set_ylabel("PnL ($)")
    ax2.legend(fontsize=9)
    ax2.grid(True, alpha=0.3)

    # ── Chart 3: Weekly Fees with warm-up marking ──
    ax3 = axes[1, 0]
    weeks = [s["week"] for s in weekly_summaries]
    company = [s["company_revenue"] for s in weekly_summaries]
    base = [s["base_pool"] for s in weekly_summaries]
    outperf = [s["outperformer_pool"] for s in weekly_summaries]

    colors_warmup = ["#ffcccc" if s["is_warmup"] else "#e74c3c" for s in weekly_summaries]
    ax3.bar(weeks, company, label="Company Revenue (33%)", color=colors_warmup, alpha=0.8)
    ax3.bar(weeks, base, bottom=company, label="Base Dividend (20%)", color="#3498db", alpha=0.8)
    ax3.bar(weeks, outperf,
            bottom=[c + b for c, b in zip(company, base)],
            label="Outperformer Pool (80%)", color="#2ecc71", alpha=0.8)

    ax3.set_title("Weekly Fee Breakdown")
    ax3.set_xlabel("Week")
    ax3.set_ylabel("Amount ($)")
    ax3.legend(fontsize=8)
    ax3.grid(True, alpha=0.3, axis="y")

    # ── Chart 4: Final PnL Distribution ──
    ax4 = axes[1, 1]
    type_pnls = {}
    for agent in agents:
        pnl = agent.portfolio_value(engine) - agent.initial_cash
        if agent.agent_type not in type_pnls:
            type_pnls[agent.agent_type] = []
        type_pnls[agent.agent_type].append(pnl)

    positions = []
    labels = []
    for i, (atype, pnls) in enumerate(type_pnls.items()):
        color = type_colors.get(atype, "#95a5a6")
        bp = ax4.boxplot(pnls, positions=[i], widths=0.6,
                         patch_artist=True,
                         boxprops=dict(facecolor=color, alpha=0.6),
                         medianprops=dict(color="black", linewidth=2))
        labels.append(atype)
        positions.append(i)

    ax4.set_xticks(positions)
    ax4.set_xticklabels(labels)
    ax4.axhline(y=0, color="red", linewidth=0.8, linestyle="--")
    ax4.set_title("Final PnL Distribution by Agent Type")
    ax4.set_ylabel("PnL ($)")
    ax4.grid(True, alpha=0.3, axis="y")

    plt.tight_layout()
    path = os.path.join(config.output_dir, "simulation_charts.png")
    fig.savefig(path, dpi=150, bbox_inches="tight")
    plt.close(fig)
    return path


# ── CLI ───────────────────────────────────────────────────────────

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Dividend Fantasy Simulation v2")
    parser.add_argument("--players", type=int, default=20, help="Number of players")
    parser.add_argument("--weeks", type=int, default=10, help="Number of weeks")
    parser.add_argument("--random-bots", type=int, default=8, help="Number of random bots")
    parser.add_argument("--momentum-bots", type=int, default=5, help="Number of momentum bots")
    parser.add_argument("--value-bots", type=int, default=3, help="Number of value bots")
    parser.add_argument("--passive-bots", type=int, default=5, help="Number of passive bots")
    parser.add_argument("--rebalancer-bots", type=int, default=4, help="Number of rebalancer bots")
    parser.add_argument("--cash", type=float, default=1000, help="Starting cash per agent")
    parser.add_argument("--top-n", type=int, default=10, help="Top N outperformers for dividends")
    parser.add_argument("--fee", type=float, default=0.015, help="Trading fee rate")
    parser.add_argument("--pool-shares", type=int, default=5000, help="Initial pool liquidity")
    parser.add_argument("--warmup", type=int, default=2, help="Warm-up weeks excluded from report")
    parser.add_argument("--churn", type=float, default=0.06, help="Weekly churn rate")
    parser.add_argument("--seed", type=int, default=None, help="Random seed")
    parser.add_argument("--output", type=str, default=None, help="Output directory")

    args = parser.parse_args()

    config = SimConfig(
        num_players=args.players,
        agent_config={
            "random": args.random_bots,
            "momentum": args.momentum_bots,
            "value": args.value_bots,
            "passive": args.passive_bots,
            "rebalancer": args.rebalancer_bots,
        },
        starting_cash=args.cash,
        num_weeks=args.weeks,
        top_n_outperformers=args.top_n,
        initial_pool_shares=args.pool_shares,
        fee_rate=args.fee,
        warmup_weeks=args.warmup,
        churn_rate=args.churn,
        seed=args.seed,
        output_dir=args.output,
    )

    run_simulation(config)
